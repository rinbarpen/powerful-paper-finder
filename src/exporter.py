"""Excel exporter — generates formatted spreadsheet of top papers."""

import logging
from datetime import datetime
from pathlib import Path
from typing import Any, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="Microsoft YaHei", size=11, bold=True, color="FFFFFF")
BODY_FONT = Font(name="Microsoft YaHei", size=10)
SCORE_HIGH_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
SCORE_MED_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
SCORE_HIGH_FONT = Font(name="Microsoft YaHei", size=10, color="006100")
SCORE_MED_FONT = Font(name="Microsoft YaHei", size=10, color="9C5700")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

FIXED_COLUMNS = [
    ("排名", 6),
    ("arXiv ID", 14),
    ("标题", 60),
    ("作者", 30),
    ("关键词", 24),
    ("领域", 16),
    ("通用", 6),
    ("通过", 10),
    ("分类", 16),
    ("发布日期", 12),
]


def _build_columns(config: dict) -> list[tuple[str, int]]:
    cols = list(FIXED_COLUMNS)
    for model in config.get("models", []):
        cols.append((f"{model['name']}评分", 14))
    cols.append(("综合得分", 10))
    cols.append(("评审理由", 50))
    return cols


def _build_columns(config: dict) -> list[tuple[str, int]]:
    cols = list(FIXED_COLUMNS)
    for model in config.get("models", []):
        cols.append((f"{model['name']}评分", 14))
    cols.append(("综合得分", 10))
    cols.append(("评审理由", 50))
    return cols


def _sorted_model_names(papers: List[dict], models_cfg: list[dict]) -> list[str]:
    return [m["name"] for m in models_cfg]


def export(papers: List[dict], config: dict) -> Path:
    output_dir = Path(config["output"]["dir"])
    output_dir.mkdir(parents=True, exist_ok=True)

    date_str = datetime.now().strftime("%Y%m%d")
    filename = config["output"]["filename_template"].format(date=date_str)
    path = output_dir / filename

    wb = Workbook()
    ws = wb.active
    ws.title = "Top Papers"

    columns = _build_columns(config)
    model_names = _sorted_model_names(papers, config.get("models", []))

    _write_header(ws, columns)
    _write_data(ws, papers, model_names, columns)
    _auto_column_width(ws, columns)

    wb.save(path)
    logger.info(f"Exported {len(papers)} papers to {path}")
    return path


def _write_header(ws, columns):
    for col, (name, _) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col, value=name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def _write_data(ws, papers, model_names, columns):
    num_model_cols = len(model_names)
    final_col = 11 + num_model_cols
    reason_col = final_col + 1

    for rank, paper in enumerate(papers, 1):
        row = rank + 1
        _cell(ws, row, 1, rank)
        _cell(ws, row, 2, paper["id"])
        _cell(ws, row, 3, paper["title"])
        _cell(ws, row, 4, ", ".join(paper["authors"][:5]) + (
            " et al." if len(paper["authors"]) > 5 else ""
        ))
        _cell(ws, row, 5, ", ".join(paper.get("keywords", [])))
        _cell(ws, row, 6, paper.get("field", ""))
        _cell(ws, row, 7, "是" if paper.get("general") else "否")
        _cell(ws, row, 8, paper.get("pass_reason", ""))
        _cell(ws, row, 9, ", ".join(paper.get("categories", [])))
        pub = paper.get("published")
        _cell(ws, row, 10, pub.strftime("%Y-%m-%d") if hasattr(pub, "strftime") else str(pub))

        scores = paper.get("model_scores", {})
        for i, name in enumerate(model_names):
            _score_cell(ws, row, 11 + i, scores.get(name))

        _score_cell(ws, row, final_col, paper.get("final_score", 0))

        reasons = paper.get("model_reasons", {})
        reason_parts = [f"[{name}]: {reasons[name]}" for name in model_names if reasons.get(name)]
        _cell(ws, row, reason_col, " | ".join(reason_parts))


def _cell(ws, row, col, value):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = BODY_FONT
    cell.border = THIN_BORDER
    cell.alignment = Alignment(vertical="center", wrap_text=True)
    return cell


def _score_cell(ws, row, col, score):
    cell = ws.cell(row=row, column=col, value=score if score is not None else "N/A")
    cell.border = THIN_BORDER
    cell.alignment = Alignment(horizontal="center", vertical="center")
    if isinstance(score, (int, float)):
        if score >= 8.5:
            cell.fill = SCORE_HIGH_FILL
            cell.font = SCORE_HIGH_FONT
        elif score >= 7.0:
            cell.fill = SCORE_MED_FILL
            cell.font = SCORE_MED_FONT
        else:
            cell.font = BODY_FONT
    else:
        cell.font = BODY_FONT
    return cell


def _auto_column_width(ws, columns):
    for col, (_, width) in enumerate(columns, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = "A2"


def export_json(papers: List[dict], output_dir: Path) -> Path:
    """Export top papers to JSON for web frontend."""
    import json

    output_dir.mkdir(parents=True, exist_ok=True)
    path = output_dir / "papers_data.json"

    def _serialize(obj: Any) -> Any:
        if hasattr(obj, "isoformat"):
            return obj.isoformat()
        return str(obj)

    data = {
        "generated_at": datetime.now().isoformat(),
        "total": len(papers),
        "papers": papers,
    }

    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2, default=_serialize)

    logger.info(f"Exported {len(papers)} papers to {path}")
    return path
